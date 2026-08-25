"""Utilities for summarizing breseq runs and building mutation comparison files.

This module builds a per-sample summary DataFrame from a batch of breseq runs
and writes/formats an Excel workbook comparing mutations across samples.
"""

import os
import re
from collections import Counter
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aisynbiopipeline.workflows.breseq import Breseq
from aisynbiopipeline.workflows.mapping import PileupCol
from aisynbiopipeline.workflows.reference_utils import (
    genomic_region_from_features,
    get_ref_genomes_path,
)
from aisynbiopipeline.limsapi.googledrive import upload_or_replace
from aisynbiopipeline.limsapi.query import query_to_dataframe


def named_group_matches(pattern: str, text: str) -> Dict[str, str | None]:
    """Return named regex group matches from text.

    Args:
        pattern: Regular expression pattern with named groups.
        text: Input text to match against.

    Returns:
        Dict where keys are named groups and values are the matched text,
        or None when a group is not present.
    """
    regex = re.compile(pattern)
    match = regex.search(text)
    if not match:
        return {name: None for name in regex.groupindex}
    return match.groupdict()


# Numeric stat fields written for every summary row (present even on error rows
# so the resulting DataFrame has a consistent set of columns).
_STAT_FIELDS = [
    'input_read_count',
    'used_read_count',
    'mapped_read_count',
    'pct_mapped',
    'consensus_mutation_count',
    'polymorphism_mutation_count',
    'average_cov',
]


def _pct(numerator, denominator):
    """Percentage rounded to one decimal, or None if either value is missing/zero."""
    if not numerator or not denominator:
        return None
    return round(numerator / denominator * 100, 1)


def _summary_error_row(title, error, sname_pattern, regions):
    """Build a summary row for a sample that could not be loaded/summarized."""
    row = {'seqsample': title}
    if sname_pattern and title:
        row.update(named_group_matches(sname_pattern, title))
    row['error'] = error
    for field in _STAT_FIELDS:
        row[field] = None
    if regions:
        for key in regions:
            row[key] = None
    return row


def create_breseq_summary(
    seqsample_batch,
    version_name,
    output_path=None,
    sname_pattern=None,
    regions=None,
    loci=None,
    gdrive_folder_id=None,
):
    """Build a per-sample breseq summary DataFrame.

    For each seqsample the corresponding breseq run is loaded and its read,
    mutation and coverage statistics are extracted. When ``sname_pattern`` is
    provided (a regex with named groups), metadata fields are parsed out of the
    seqsample name and added as columns. Optional per-region average coverage and
    per-locus allele frequencies are added when ``regions`` / ``loci`` are given.

    Args:
        seqsample_batch: Iterable of seqsample objects (need ``library.path`` and
            ``sample_name``).
        version_name: Breseq version/parameter folder name.
        output_path: If set, write the summary to this CSV path.
        sname_pattern: Regex with named groups used to parse the seqsample name.
        regions: Mapping of column name -> region string for coverage lookups.
        loci: Mapping ``{'mutations': {name: {'locus': ..., 'alt_allele': ...}}}``.
        gdrive_folder_id: If set (with ``output_path``), upload the CSV to Drive.

    Returns:
        pandas.DataFrame with one row per seqsample.
    """
    rows = []

    for s in seqsample_batch:
        breseq_folder = s.library.path / 'breseq' / s.sample_name / version_name
        title = s.sample_name

        try:
            b = Breseq.from_existing(breseq_folder)
        except Exception as e:
            print(f"Failed to load breseq run for {title}: {e}")
            rows.append(_summary_error_row(title, str(e), sname_pattern, regions))
            continue

        title = getattr(b, 'title', None) or s.sample_name

        try:
            b.count_reads()
            b.count_mutations()

            row = {'seqsample': title}
            if sname_pattern:
                row.update(named_group_matches(sname_pattern, title))
            row['error'] = None
            row['input_read_count'] = getattr(b, 'input_read_count', None)
            row['used_read_count'] = getattr(b, 'used_read_count', None)
            row['mapped_read_count'] = getattr(b, 'mapped_read_count', None)
            row['pct_mapped'] = _pct(row['mapped_read_count'], row['input_read_count'])
            row['consensus_mutation_count'] = getattr(b, 'consensus_mutation_count', None)
            row['polymorphism_mutation_count'] = getattr(b, 'polymorphism_mutation_count', None)
            row['average_cov'] = getattr(b, 'avg_coverage', None)

            if regions:
                for key, region in regions.items():
                    row[key] = b.get_region_average_coverage(region)

            if loci:
                for key, value in loci['mutations'].items():
                    basecalls = PileupCol(b.bam_path, value['locus']).basecalls
                    locus_cov = len(basecalls)
                    counts = Counter(basecalls)
                    alt_freq = counts[value['alt_allele']] / locus_cov if locus_cov else None
                    row[key + '_locus_cov'] = locus_cov
                    row[key + '_alleleCounts'] = dict(counts)
                    row[key + '_alt_allele_freq'] = alt_freq
        except Exception as e:
            print(f"Error summarizing breseq run for {title}: {e}")
            rows.append(_summary_error_row(title, str(e), sname_pattern, regions))
            continue

        rows.append(row)

    breseq_summary = pd.DataFrame(rows)

    # Copy-number estimate per region = region coverage / genome-wide coverage.
    if regions:
        for key in regions:
            breseq_summary[key + '_CN'] = breseq_summary[key] / breseq_summary['average_cov']

    breseq_summary['is_parent'] = breseq_summary['seqsample'].apply(
        lambda x: bool(x) and x.startswith('ANL')
    )

    # Put metadata columns first, then everything else. Only include metadata
    # columns that actually exist (they only appear when sname_pattern parsed them).
    metadata_cols = [
        'seqsample',
        'Experiment',
        'Plate_name',
        'Strain',
        'DNA_construct',
        'Replicate',
        'Transfer',
        'Isolate',
        'Colony_number',
    ]
    metadata_cols = [c for c in metadata_cols if c in breseq_summary.columns]
    other_cols = [c for c in breseq_summary.columns if c not in metadata_cols]
    breseq_summary = breseq_summary[metadata_cols + other_cols]

    # Sort by whichever of the preferred sort keys are present. Parents first
    # (is_parent descending), remaining keys ascending.
    preferred_sort = ['is_parent', 'DNA_construct', 'Strain', 'Replicate', 'Transfer']
    sort_cols = [c for c in preferred_sort if c in breseq_summary.columns]
    if sort_cols:
        breseq_summary.sort_values(
            sort_cols,
            ascending=[c != 'is_parent' for c in sort_cols],
            inplace=True,
        )

    # Write breseq run summary to csv
    if output_path:
        breseq_summary.to_csv(output_path, index=False)

        # Upload to Google Drive (requires the CSV to have been written)
        if gdrive_folder_id:
            upload_or_replace(output_path, gdrive_folder_id)

    return breseq_summary


def _parse_region(region):
    """Split a ``"genome:start-stop"`` region string into its parts.

    Returns a tuple ``(genome, start, stop)`` where ``start``/``stop`` are ints,
    or ``(None, None, None)`` if the string cannot be parsed.
    """
    try:
        genome, coords = region.rsplit(':', 1)
        start_str, stop_str = coords.split('-')
        return genome, int(start_str), int(stop_str)
    except (ValueError, AttributeError):
        return None, None, None


# Column order expected by aisynbiopipeline/pipeline/db_update_copy_numbers.py.
_COPY_NUMBER_COLUMNS = [
    'Seqsample',
    'Seqorder',
    'Breseq_registry_ID',
    'Refgenome',
    'Refgenome_avg_cov',
    'Region_name',
    'Region_start',
    'Region_stop',
    'Region_avg_cov',
    'Region_CN',
]


def create_copy_number_csv(
    breseq_batch,
    regions,
    output_path=None,
    gdrive_folder_id=None,
):
    """Build a copy-number table (one row per breseq run x region).

    Produces a DataFrame/CSV in the format consumed by
    ``aisynbiopipeline/pipeline/db_update_copy_numbers.py``. For each breseq run
    and each region, the region's average coverage is compared to the run's
    genome-wide average coverage to give a copy-number estimate.

    Args:
        breseq_batch: Iterable of :class:`~aisynbiopipeline.workflows.breseq.Breseq`
            objects (loaded via ``Breseq.from_existing``).
        regions: Mapping of region name -> region string ``"genome:start-stop"``,
            exactly like the ``regions`` argument of :func:`create_breseq_summary`.
        output_path: If set, write the table to this CSV path.
        gdrive_folder_id: If set (with ``output_path``), upload the CSV to Drive.

    Returns:
        pandas.DataFrame with one row per (breseq run, region), with the columns
        expected by the Copy_numbers DB update script.
    """
    rows = []

    for b in breseq_batch:
        seqsample = getattr(b, 'title', None)

        # Breseq_registry_ID is the on-disk version folder name (authoritative).
        try:
            output_folder = Path(b.output_folder)
            breseq_registry_id = output_folder.name
        except (AttributeError, TypeError):
            output_folder = None
            breseq_registry_id = None

        # Seqorder is the seqorder folder in the standard run path layout:
        # <seqorder>/<seqorder>_<platform>/breseq/<sample>/<version>
        seqorder = None
        if output_folder is not None:
            try:
                seqorder = output_folder.parents[3].name
            except IndexError:
                seqorder = None

        refgenome_avg_cov = getattr(b, 'avg_coverage', None)

        for region_name, region in regions.items():
            genome, start, stop = _parse_region(region)

            try:
                region_avg_cov = b.get_region_average_coverage(region)
            except Exception as e:
                print(f"Failed to get coverage for {seqsample} region {region_name}: {e}")
                region_avg_cov = None

            if region_avg_cov is not None and refgenome_avg_cov:
                region_cn = region_avg_cov / refgenome_avg_cov
            else:
                region_cn = None

            rows.append({
                'Seqsample': seqsample,
                'Seqorder': seqorder,
                'Breseq_registry_ID': breseq_registry_id,
                'Refgenome': genome,
                'Refgenome_avg_cov': refgenome_avg_cov,
                'Region_name': region_name,
                'Region_start': start,
                'Region_stop': stop,
                'Region_avg_cov': region_avg_cov,
                'Region_CN': region_cn,
            })

    copy_number_df = pd.DataFrame(rows, columns=_COPY_NUMBER_COLUMNS)

    if output_path:
        copy_number_df.to_csv(output_path, index=False)

        # Upload to Google Drive (requires the CSV to have been written)
        if gdrive_folder_id:
            upload_or_replace(output_path, gdrive_folder_id)

    return copy_number_df


def get_region_parameter(genbank_file, feature_first, feature_last):
    genbank_file_path = os.path.join(get_ref_genomes_path(), genbank_file)
    genome, start, stop = genomic_region_from_features(genbank_file_path, feature_first, feature_last)
    region = genome + ":" + str(start) + "-" + str(stop)
    return region


def reformat_comparison_df(comparison_csv, breseq_summary):
    """Pivot a gdtools comparison CSV into a mutation-by-sample frequency table."""
    compare_df = pd.read_csv(comparison_csv)

    # Need to exclude these because they will mess with pivoting on title
    cols_with_diff_vals_for_same_mutation = [
        'new_read_count', 'new_read_count_basis', 'ref_read_count',
        'ref_read_count_basis', 'multiple_polymorphic_SNPs_in_same_codon',
        'repeat_new_copies', 'repeat_ref_copies',
    ]

    # Exclude these because they are not of interest
    cols_uninformative = ['clone', 'mutator_status', 'population', 'time', 'treatment', 'transl_table']

    cols_to_drop = set(cols_with_diff_vals_for_same_mutation) | set(cols_uninformative)
    cols_to_keep = [col for col in compare_df.columns if col not in cols_to_drop]

    # Columns that should only have integer values. Some may have weird text
    # which messes with conversion to nullable int type, so force to numeric.
    int_cols = [
        'aa_position',
        'codon_number',
        'codon_position',
        'gene_position',
        'insert_position',
        'position',
        'position_end',
        'position_start',
        'repeat_length',
        'size',
    ]
    for col in int_cols:
        if col in compare_df.columns:
            compare_df[col] = pd.to_numeric(compare_df[col], errors='coerce')

    # Now that expected numeric cols are numeric, can convert to int if possible
    compare_df = compare_df[cols_to_keep].convert_dtypes()

    # Define columns that will form the MultiIndex
    index = [col for col in compare_df.columns if col not in ('title', 'frequency')]

    # Create the pivoted mutation frame, where distinct mutations are rows,
    # samples are columns, values are frequencies.
    df = compare_df.pivot(index=index, columns='title', values='frequency').sort_index(level='position')
    df = df.fillna(0)

    # Order the comparison just like the summary
    df = df[breseq_summary['seqsample'].to_list()]

    return df


# ------ Make highlight color dictionary ------

def create_highlight_color_dict(
    index_col_names,
    breseq_summary,
    highlight_parents=False,
    highlight_alt_samples=False,
):
    """Map each workbook column header to an openpyxl fill.

    Index (metadata) columns are gray; sample columns default to white. Optional
    highlighting groups ALE transfer samples by their originating sample in
    alternating colors and/or highlights parent samples green.
    """
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    green_fill = PatternFill(start_color="E5FFCC", end_color="E5FFCC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

    highlight_colors = {}
    highlight_colors.update(dict(zip(index_col_names, [gray_fill] * len(index_col_names))))

    seqsamples = breseq_summary['seqsample'].to_list()
    highlight_colors.update(dict(zip(seqsamples, [white_fill] * len(seqsamples))))

    if highlight_alt_samples and 'sample' in breseq_summary.columns:
        # ALE transfer samples grouped by original sample, alternating colors.
        unique_samples = breseq_summary['sample'].unique()
        alt_colors = dict(zip(unique_samples, len(unique_samples) * [white_fill, orange_fill]))
        ale_samples = breseq_summary.loc[~breseq_summary['is_parent']]
        highlight_colors.update(
            dict(zip(
                ale_samples['seqsample'].to_list(),
                ale_samples['sample'].apply(lambda x: alt_colors[x]).to_list(),
            ))
        )

    if highlight_parents:
        parents = breseq_summary.loc[breseq_summary['is_parent']]['seqsample'].to_list()
        highlight_colors.update(dict(zip(parents, [green_fill] * len(parents))))

    return highlight_colors


# ------ Build workbook (all samples, individual samples) ------

def write_mutation_comparison_wb(
    df,
    breseq_summary,
    excel_path,
    add_per_sample_sheets=False,
    sample_col='sample',
):
    """Write the mutation comparison workbook.

    The first sheet ("all_samples") contains every sample column. When
    ``add_per_sample_sheets`` is True, additional sheets are written per group in
    ``breseq_summary[sample_col]``, each showing that group's parent + transfers.
    """
    mutation_df = df.copy()

    def get_parent_sample(sample):
        lims_samples = query_to_dataframe('Samples')[['Name', 'Parent_sample']]
        lims_seqsamples = query_to_dataframe('Seqsamples')[['Sequencing_sample', 'Sample_Name']]
        parent_samples = pd.merge(
            lims_seqsamples,
            lims_samples,
            left_on='Sample_Name',
            right_on="Name",
            how="left",
        ).drop_duplicates()
        return parent_samples.loc[parent_samples['Sample_Name'] == sample]['Parent_sample'].iloc[0]

    # Excel sheet names can only have 31 chars, which is much less than many of
    # the seqsample names. Name per-sample sheets by their Excel column range
    # (A, B, C, ... AA, AB, ...) instead.
    df_col_names = df.reset_index().columns
    excel_col_names = [get_column_letter(i) for i in range(1, len(df_col_names) + 1)]
    excel_col_name_dict = dict(zip(df_col_names, excel_col_names))

    def get_sheet_name(seqsamples):
        letters = sorted(excel_col_name_dict[x] for x in seqsamples if x in excel_col_name_dict)
        return f"{letters[0]} to {letters[-1]}"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # First sheet contains all columns
        mutation_df.reset_index().to_excel(writer, sheet_name="all_samples", index=False)

        if add_per_sample_sheets and sample_col in breseq_summary.columns:
            for name, group in breseq_summary.groupby(sample_col, sort=False):
                if name == 'NA':
                    continue
                try:
                    parent_sample = get_parent_sample(name)
                    cols = [parent_sample] + group['seqsample'].to_list()
                except Exception:
                    cols = group['seqsample'].to_list()

                cols = [c for c in cols if c in mutation_df.columns]
                trimmed_df = mutation_df[cols]
                # Discard mutations that are zero in all of this group's samples
                trimmed_df = trimmed_df.loc[~(trimmed_df == 0).all(axis=1)]
                sheet_name = get_sheet_name(group['seqsample'].to_list())
                trimmed_df.reset_index().to_excel(writer, sheet_name=sheet_name, index=False)


# ------ Format workbook: format header, format columns ------

def format_mutation_comparison_wb(excel_path, highlight_colors):
    wb = load_workbook(excel_path)

    # Border formatting
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    for ws in wb.worksheets:
        header_row = ws[1]

        # Header row formatting
        header_font = Font(bold=True)
        header_alignment = Alignment(wrap_text=True)
        for cell in header_row:
            cell.font = header_font
            cell.alignment = header_alignment
        ws.freeze_panes = "A2"

        # Highlight columns by header name
        for col_index, header_cell in enumerate(header_row):
            fill = highlight_colors.get(header_cell.value)
            if fill is None:
                continue
            for row_cells in ws.iter_rows(min_row=1):
                row_cells[col_index].fill = fill

    wb.save(excel_path)


def write_and_format_mutation_comparison_excel(
    excel_path,
    mutation_comparison_df,
    breseq_summary,
    gdrive_folder_id=None,
):
    # Index (metadata) columns are the MultiIndex level names of the pivoted
    # comparison frame; the data columns are the sample names.
    index_col_names = [n for n in mutation_comparison_df.index.names if n not in (None, 'title', 'frequency')]

    print("Assigning fill colors to columns.")
    highlight_colors = create_highlight_color_dict(index_col_names, breseq_summary, highlight_alt_samples=True)
    print("\tDone.")
    print("Writing mutation comparison to Excel workbook.")
    write_mutation_comparison_wb(mutation_comparison_df, breseq_summary, excel_path)
    print("\tDone.")
    print("Formatting Excel workbook.")
    format_mutation_comparison_wb(excel_path, highlight_colors)
    if gdrive_folder_id:
        print("Uploading Excel workbook to Google Drive.")
        upload_or_replace(excel_path, gdrive_folder_id)
    print("\tDone.")
